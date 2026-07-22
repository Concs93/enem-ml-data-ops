"""
Gera os seeds do dbt a partir dos artefatos oficiais do INEP.

Entrada : a pasta dos microdados descompactados (scripts INPUT_R_*.R e o
          dicionario .xlsx).
Saida   : dbt/seeds/dominios.csv e dbt/seeds/co_prova.csv

Nenhum CSV e mantido a mao: baixou os microdados, roda isto e os seeds sao
derivados da fonte oficial.

Uso:
    python -m ingestion.build_seeds --microdados "data/raw/microdados_enem_2025"
"""

import argparse
import csv
import glob
import os
import re
import sys


# ---------------------------------------------------------------- utilidades

def _split_args(s):
    """Divide 'a', 'b, c', 3 respeitando aspas."""
    out, buf, quote = [], "", None
    for ch in s:
        if quote:
            if ch == quote:
                quote = None
            else:
                buf += ch
        elif ch in "'\"":
            quote = ch
        elif ch == ",":
            out.append(buf.strip())
            buf = ""
        else:
            buf += ch
    if buf.strip() or out:
        out.append(buf.strip())
    return [x for x in out if x != ""]


def _extrai_c(corpo, palavra):
    """Extrai o conteudo de `palavra = c(...)` contando parenteses."""
    m = re.search(palavra + r"\s*=\s*c\(", corpo)
    if not m:
        return None
    i, prof = m.end(), 1
    while i < len(corpo) and prof:
        if corpo[i] == "(":
            prof += 1
        elif corpo[i] == ")":
            prof -= 1
        i += 1
    return corpo[m.end(): i - 1]


# --------------------------------------------------- fonte 1: scripts INPUT_R

def dominios_dos_scripts_r(pasta):
    """Le os INPUT_R_*.R e devolve {(base, variavel): [(codigo, rotulo)]}."""
    arquivos = sorted(glob.glob(os.path.join(pasta, "**", "INPUT_R_*.R"),
                                recursive=True))
    if not arquivos:
        raise SystemExit(
            f"Nenhum INPUT_R_*.R encontrado em {pasta}.\n"
            "Aponte --microdados para a pasta dos microdados descompactados."
        )

    dominios, divergencias = {}, []
    for caminho in arquivos:
        nome = os.path.basename(caminho)
        m = re.search(r"INPUT_R_(.+?)_\d{4}\.R$", nome)
        base = m.group(1).lower() if m else nome.lower()

        txt = open(caminho, encoding="utf-8", errors="replace").read()
        txt = txt.replace("\r", "")
        # as linhas de rotulo vem comentadas; removemos apenas o marcador
        txt = "\n".join(re.sub(r"^\s*#", "", ln) for ln in txt.splitlines())

        for m in re.finditer(r"\$([A-Z0-9_]+)\s*<-\s*factor\(", txt):
            var, i, prof = m.group(1), m.end(), 1
            while i < len(txt) and prof:
                if txt[i] == "(":
                    prof += 1
                elif txt[i] == ")":
                    prof -= 1
                i += 1
            corpo = txt[m.end(): i - 1]

            niveis = _extrai_c(corpo, "levels")
            rotulos = _extrai_c(corpo, "labels")
            if not (niveis and rotulos):
                continue
            niveis, rotulos = _split_args(niveis), _split_args(rotulos)

            if len(niveis) != len(rotulos):
                divergencias.append((base, var, len(niveis), len(rotulos)))
                continue
            dominios[(base, var)] = list(zip(niveis, rotulos))

        print(f"  lido {nome}")
    return dominios, divergencias


# -------------------------------------------------- fonte 2: dicionario .xlsx

def dominios_do_dicionario(pasta):
    """Le o dicionario .xlsx. E a fonte autoritativa em caso de conflito."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("  ! openpyxl nao instalado; pulando validacao pelo dicionario")
        return {}

    candidatos = [p for p in glob.glob(os.path.join(pasta, "**", "*.xlsx"),
                                       recursive=True)
                  if "dicion" in os.path.basename(p).lower()]
    if not candidatos:
        print("  ! dicionario .xlsx nao encontrado; pulando validacao")
        return {}

    wb = load_workbook(candidatos[0], read_only=True)
    print(f"  lido {os.path.basename(candidatos[0])}")

    dominios = {}
    for aba in wb.sheetnames:
        base = re.sub(r"_\d{4}$", "", aba).lower()
        var_atual, pares = None, []

        for linha in wb[aba].iter_rows(values_only=True):
            celulas = [("" if c is None else str(c).strip()) for c in linha]
            celulas += [""] * (6 - len(celulas))
            primeira = celulas[0]

            # o par codigo/rotulo mora nas colunas 3 e 4, tanto na linha
            # que abre a variavel quanto nas linhas de continuacao
            cod, rot = celulas[2], celulas[3]
            valido = bool(cod) and bool(rot) and len(cod) <= 3

            # inicio de variavel: NOME | descricao | codigo | rotulo | ...
            if re.fullmatch(r"[A-Z][A-Z0-9_]{2,}", primeira):
                if var_atual and pares:
                    dominios[(base, var_atual)] = pares
                var_atual, pares = primeira, []
                if valido:
                    pares.append((cod, rot))
                continue

            # continuacao: as duas primeiras colunas vem vazias
            if var_atual and not primeira and valido:
                pares.append((cod, rot))

        if var_atual and pares:
            dominios[(base, var_atual)] = pares

    return dominios


# ------------------------------------------------------------ classificacao

def classifica_prova(rotulo):
    """Regra duravel: e regular quando o rotulo e so a cor, sem qualificador."""
    if "BAM" in rotulo:
        return "bam"
    if "Reaplica" in rotulo:
        return "reaplicacao"
    if " - " in rotulo:
        return "acessibilidade"
    return "regular"


# ------------------------------------------------------------------- escrita

def escreve(caminho, cabecalho, linhas):
    os.makedirs(os.path.dirname(caminho), exist_ok=True)
    with open(caminho, "w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(cabecalho)
        w.writerows(linhas)
    print(f"  gravado {caminho} ({len(linhas)} linhas)")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--microdados", required=True,
                    help="pasta dos microdados descompactados")
    ap.add_argument("--saida", default="dbt/seeds",
                    help="pasta de destino dos seeds (padrao: dbt/seeds)")
    args = ap.parse_args()

    print("1. Scripts INPUT_R")
    do_r, divergencias = dominios_dos_scripts_r(args.microdados)

    print("2. Dicionario oficial")
    do_dic = dominios_do_dicionario(args.microdados)

    print("3. Reconciliacao")
    conflitos, suspeitos = 0, 0
    dominios = dict(do_r)
    for chave, pares in do_dic.items():
        if chave not in dominios:
            dominios[chave] = pares
            continue
        atual = dominios[chave]
        if atual == pares:
            continue
        if len(pares) > len(atual):
            # dicionario mais completo: ele e a fonte autoritativa
            print(f"  ! {chave[1]}: script R tem {len(atual)} codigos, "
                  f"dicionario tem {len(pares)} -> usando dicionario")
            dominios[chave] = pares
            conflitos += 1
        elif len(pares) < len(atual):
            # dicionario menos completo costuma indicar falha de leitura
            # da planilha, nao divergencia real: mantemos o script R
            print(f"  ~ {chave[1]}: dicionario leu menos codigos que o script R "
                  f"({len(pares)} x {len(atual)}) -> mantendo script R")
            suspeitos += 1
        else:
            dominios[chave] = pares

    for base, var, n_niv, n_rot in divergencias:
        if (base, var) in dominios:
            print(f"  + {var}: script R inconsistente ({n_niv} niveis x "
                  f"{n_rot} rotulos), resolvido pelo dicionario")
        else:
            print(f"  ! {var}: script R inconsistente e ausente do dicionario "
                  f"-> NAO mapeado")

    # seed: dominios
    linhas = []
    for (base, var), pares in sorted(dominios.items()):
        for cod, rot in pares:
            linhas.append([base, var, cod, rot])
    escreve(os.path.join(args.saida, "dominios.csv"),
            ["base", "variavel", "codigo", "rotulo"], linhas)

    # seed: co_prova
    provas = []
    for (base, var), pares in dominios.items():
        m = re.fullmatch(r"CO_PROVA_(CN|CH|LC|MT)", var)
        if not m:
            continue
        area = m.group(1)
        for cod, rot in pares:
            tipo = classifica_prova(rot)
            cor = rot.split(" - ")[0].split(" (")[0].strip()
            provas.append([cod, area, cor, rot, tipo,
                           "true" if tipo == "regular" else "false"])
    provas.sort(key=lambda r: (r[1], int(r[0])))
    escreve(os.path.join(args.saida, "co_prova.csv"),
            ["co_prova", "sg_area", "cor", "rotulo", "tipo_aplicacao",
             "is_regular"], provas)

    # resumo
    regulares = [p for p in provas if p[5] == "true"]
    print("\nResumo")
    print(f"  variaveis mapeadas : {len(dominios)}")
    print(f"  versoes de prova   : {len(provas)}")
    print(f"  conflitos resolvidos pelo dicionario : {conflitos}")
    print(f"  leituras suspeitas do dicionario     : {suspeitos}")
    print("  provas regulares por area:")
    for area in ["CH", "CN", "LC", "MT"]:
        reg = [p for p in regulares if p[1] == area]
        print(f"    {area}: {', '.join(p[0] for p in reg)}  "
              f"({', '.join(p[2] for p in reg)})")

    if not regulares:
        print("\n! Nenhuma prova regular identificada - verifique os rotulos.")
        sys.exit(1)


if __name__ == "__main__":
    main()